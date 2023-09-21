from django.shortcuts import render, redirect, reverse
from django.contrib import messages
from django.contrib.auth.models import User, auth
from .models import *
from .forms import *
from .decorators import *
import openpyxl
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .filters import HardwareFilter
from django_filters.views import FilterView
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Q

PRODUCTS_PER_PAGE = 25


##Funciones

def mayus_minus(pal):
    if(pal == ''):
        return pal
    pal.strip()
    if pal[-1] == ' ':
        pal.replace(' ', '')
    return pal.lower().capitalize()

def buscar_repetido(nro_de_serie):
    try:
        hard = Hardware.objects.get(nro_de_serie = nro_de_serie)
    except:
        hard = None
    return hard
# Create your views here.

@login_required(login_url='login')
def index(request):
    cant_pags = []
    page = request.GET.get('page',1)
    asignacion = request.GET.get('asignacion', False)
    editar = request.GET.get('editar', False)
    agregar = request.GET.get('agregar', False)
    
    
    f = HardwareFilter(request.GET, queryset=Hardware.objects.all())
    
    product_paginator = Paginator(list(f.qs), PRODUCTS_PER_PAGE)
    
    try:
        pagina = product_paginator.page(page)
    except EmptyPage:
        pagina = product_paginator.page(product_paginator.num_pages)
    except:
        pagina = product_paginator.page(product_paginator.num_pages)

    for i in list(product_paginator.page_range):
        if i <= int(page) + 5 and i>=int(page)-5:
            cant_pags.append(i)

    ctx = {
        'link':'index',
        'filter':f,
        'pagina': pagina,
        'paginator':product_paginator,
        'cant_pags':cant_pags,
        'num_page':int(page)
    }

    if asignacion == '1':
        ctx['asignacion'] = True
    if editar == '1':
        ctx['editar'] = True
    if agregar == '1':
        ctx['agregar'] = True
    
    
    return render(request, 'main.html', ctx)

@unauthorized_user
def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password2 = request.POST['password2']
        name = request.POST['name']
        surname = request.POST['surname']
        
        

        if password == password2:
            if User.objects.filter(email=email).exists():
                messages.info(request, 'Email ya usado')
                return redirect('register')
            elif User.objects.filter(username=username).exists():
                messages.info(request, 'Username Taken')
                return redirect('register')
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                user.save()
                group, create=Group.objects.get_or_create(name='user')
                user_model = User.objects.get(username=username)
                user_model.groups.add(group)
                nuevo_tecnico = Tecnico.objects.create(user = user_model, id_user = user_model.id, nombre=name, apellido = surname)
                nuevo_tecnico.save()
                return redirect('login')
        else:
            messages.info(request, 'Las contraseñas no coinciden.')
            return redirect('register') 
    return render(request,'signup.html')

@unauthorized_user
def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        

        user = auth.authenticate(username=username, password=password)
        

        if user is not None:
            auth.login(request, user)
            return redirect('/')
        else:
            messages.info(request,'Credentials invalid')
            return redirect('login')
    return render(request, 'signin.html')

@login_required(login_url='login')
def logout(request):
    auth.logout(request)
    return redirect('login')

@login_required(login_url='login')
def add_inventary(request):
    
    form_add_inventary = HardwareForm()
    
    ctx = {'link':'create'}
    ctx['form_add_inventary'] = form_add_inventary

    if request.method == 'POST':
        form = HardwareForm(request.POST)
        if form.is_valid():
            try:
                form.cleaned_data['nro_de_serie'] = form.cleaned_data['nro_de_serie'].upper()
            except:
                pass
            if len(form.cleaned_data['nro_de_serie']) > 0 and ('?' not in form.cleaned_data['nro_de_serie']):
                hard = buscar_repetido(form.cleaned_data['nro_de_serie'])
            else:
                hard = None
            
            if hard != None:
                messages.info(request, 'Ya hay un hardware en el inventario con el mismo numero de serie.')
                ctx['form_add_inventary'] = HardwareForm(form.cleaned_data)
            else:
                print(form.cleaned_data['nro_de_serie'])
                if len(form.cleaned_data['nro_de_serie']) == 0 or ('?' in form.cleaned_data['nro_de_serie']):
                    form.cleaned_data['nro_de_serie'] = 'S/D'
                tipo, create = Tipo.objects.get_or_create(id = request.POST['tipo'])
                marca, create = Marca.objects.get_or_create(id = request.POST['marca']) 
                modelo = Modelo.objects.get(id = request.POST['modelo'])
                ubicacion, create = Ubicacion.objects.get_or_create(id=request.POST['ubicacion'])
                estado, create = Estado.objects.get_or_create(id = request.POST['estado'])
                
                hardware = Hardware.objects.create(tipo = tipo, marca=marca, modelo=modelo,ubicacion=ubicacion, estado = estado, nro_de_serie=form.cleaned_data['nro_de_serie'], observaciones = request.POST['observaciones'])
                
                if(request.user.is_staff == False):
                    Notificacion.objects.create(hardware=hardware, usuario = request.user, tipo = 'CREATE')
                
                return redirect(reverse('index')+f'?agregar=1')
    return render(request, 'main.html', ctx)

@login_required(login_url='login')
@admin_only
def reload(request):
    df = openpyxl.load_workbook("INVENTARIO_YENNY.xlsx")
    print('yenny')
    dataframe = df.active
    data = []
    for row in range(1, dataframe.max_row):
        _row=[row]
        for col in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(col[row].value)
        data.append(_row)

    for dato in data:
            print(dato)
            tipo, create = Tipo.objects.get_or_create(name = "Pantalla")
            marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[1]))) 
            modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[2])))
            ubicacion, create = Ubicacion.objects.get_or_create(nombre=mayus_minus(str(dato[6])))
            estado, create = Estado.objects.get_or_create(nombre = 'Activo')

            if dato[7] == None:
                dato[7] = ''
            else:
                dato[7] = mayus_minus(str(dato[7]))

            hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], origen = 'Yenny')
            hard.save()

    df = openpyxl.load_workbook("Mueble Cau.xlsx")
    print('mueble cau')
    
    for i in df.sheetnames:
        
        if df[i].title == 'Asignaciones':
            pass
        else:
            dataframe = df[i]
            data = []
            for row in range(1, dataframe.max_row):
                _row=[row]
                for col in dataframe.iter_cols(1,dataframe.max_column):
                    _row.append(col[row].value)
                data.append(_row)
            
            for dato in data:
                    print(dato)
                    tipo, create = Tipo.objects.get_or_create(name = mayus_minus(str(dato[1])))
                    marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[2]))) 
                    modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[3])))
                    ubicacion, create = Ubicacion.objects.get_or_create(nombre='Mueble CAU')
                    estado, create = Estado.objects.get_or_create(nombre = 'Activo')

                    if dato[7] == None:
                        dato[7] = ''
                    else:
                        dato[7] = mayus_minus(str(dato[7]))
                    if dato[4] == None:
                        dato[4] = 'S/D'

                    hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7] + '\n' + str(dato[8]))
                    hard.save()

    df = openpyxl.load_workbook("INVENTARIO T4.xlsx")
    print('inv t4')
    for i in df.sheetnames:
        if df[i].title == 'Scrap':
            pass
        else:
            dataframe = df[i]
            data = []
            print('entre', data, dataframe)
            for row in range(1, dataframe.max_row):
                _row=[row]
                for col in dataframe.iter_cols(1,dataframe.max_column):
                    _row.append(col[row].value)
                data.append(_row)

            for dato in data:
                print(dato)
                tipo, create = Tipo.objects.get_or_create(name = mayus_minus(str(dato[1])))
                marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[2]))) 
                modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[3])))
                ubicacion, create = Ubicacion.objects.get_or_create(nombre=mayus_minus(str(dato[6])))
                estado, create = Estado.objects.get_or_create(nombre = 'Activo')

                if dato[7] == None:
                    dato[7] = ''
                else:
                    if len(dato) > 8:
                        dato[7] = mayus_minus(str(dato[7]))
                    else:
                        dato[7] = mayus_minus(str(dato[7]))
                hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], origen = "T4")
                hard.save()

    return redirect('index')

@login_required(login_url='login')
def delete(request, id):
    if(request.user.is_staff):
        hardware = Hardware.objects.get(id=id)
        hardware.delete()
        return redirect('index')
    else:
        hardware = Hardware.objects.get(id=id)
        usuario = User.objects.get(username = request.user)
        Notificacion.objects.create(hardware=hardware, usuario = usuario, tipo = 'DELETE')
        return redirect('index')

@login_required(login_url='login')
def edit(request, id):
    try:
        to_edit = Hardware.objects.get(id=id)
    except:
        return redirect('index')
    ctx = {}
    ctx['to_edit'] = to_edit
    edit_form = HardwareForm(to_edit.toJSON())
    ctx['edit_form'] = edit_form
    ctx['link'] = 'edit'
    
    
    if request.method == 'POST':
        
        if request.POST['nro_de_serie'].upper() != to_edit.nro_de_serie:
            
            if len(request.POST['nro_de_serie']) > 0 and ('?' not in request.POST['nro_de_serie']):
                hard = Hardware.objects.filter(nro_de_serie = request.POST['nro_de_serie'].upper())
                
            else:
                hard = False

            if hard:
                messages.info(request, 'Ya hay un hardware en el inventario con el mismo numero de serie.')
                ctx['form_add_inventary'] = HardwareForm(request.POST)
                return render(request, 'main.html', ctx)
        
        to_edit.tipo  = Tipo.objects.get(id=request.POST['tipo'])
        to_edit.marca = Marca.objects.get(id=request.POST['marca'])
        to_edit.modelo = Modelo.objects.get(id=request.POST['modelo'])
        to_edit.ubicacion = Ubicacion.objects.get(id=request.POST['ubicacion'])
        to_edit.observaciones = request.POST['observaciones']
        if(request.user.is_staff):

            to_edit.nro_de_serie = request.POST['nro_de_serie'].upper()
            to_edit.estado = Estado.objects.get(id=request.POST['estado'])
        else:
            nro_serie = ''
            estado = ''
            if to_edit.nro_de_serie != request.POST['nro_de_serie']:
                nro_serie= request.POST['nro_de_serie'].upper()
            if to_edit.estado != Estado.objects.get(id=request.POST['estado']):
                estado= (Estado.objects.get(id=request.POST['estado'])).nombre
    
            Notificacion.objects.create(hardware=to_edit, usuario = User.objects.get(username = request.user), tipo = 'EDIT', nro_de_serie = nro_serie, estado = estado)
        to_edit.save()
        
        return redirect(reverse('index')+f'?editar=1')
    
    return render(request, 'main.html', ctx)

# def test(request):
#     page = request.GET.get('page',1)
    
#     f = HardwareFilter(request.GET, queryset=Hardware.objects.all())
#     product_paginator = Paginator(list(f.qs), PRODUCTS_PER_PAGE)
#     try:
#         pagina = product_paginator.page(page)
#     except EmptyPage:
#         print('hola')
#         pagina = product_paginator.page('1')
#     ctx = {
#         'link':'test',
#         'filter':f,
#         'pagina': pagina,
#         'paginator':product_paginator
#     }
#     return render(request, 'main.html', ctx)

def get_info(request):
    data = list(Hardware.objects.values())
    
    return JsonResponse(data, safe=False)

@login_required(login_url='login')
@admin_only
def notificaciones(request):
    ctx={'link':'notification'}
    status = request.GET.get('status', False)
    

    if status == 'cancel':
        ctx['status'] = True
        ctx['title'] = 'Peticion cancelada'
        ctx['msg'] = 'Se ha cancelado la petición correctamente.'
    elif status == 'accept':
        ctx['status'] = True
        ctx['title'] = 'Peticion aprobada'
        ctx['msg'] = 'Se ha aprobado la petición correctamente.'
    
    notificaciones = Notificacion.objects.filter(realizado = False)
    
    ctx['notificaciones'] = notificaciones
    ctx['cant_notificaciones'] = len(notificaciones)
    
    return render(request, 'main.html', ctx)

@login_required(login_url='login')
@admin_only
def accion_notificacion(request):
    id = request.GET.get('id')
    status = request.GET.get('status')
    notificacion = Notificacion.objects.get(id=id)

    
    if (notificacion.tipo == 'CREATE' and status == 'cancel') or (notificacion.tipo == 'DELETE' and status == 'accept'):
        hardware = Hardware.objects.get(id = notificacion.hardware.id)
        hardware.delete()
    elif(notificacion.tipo == 'EDIT' and status == 'accept'):
        hardware = Hardware.objects.get(id = notificacion.hardware.id)
        if notificacion.nro_de_serie:
            hardware.nro_de_serie = notificacion.nro_de_serie
        if notificacion.estado:
            hardware.estado = Estado.objects.get(nombre = notificacion.estado)
        hardware.save()

    notificacion.realizado = True
    notificacion.save()

    return redirect(reverse('notifications') + f'?status={status}')

def asignacion(request):
    id = request.GET.get('id')
    hardware = Hardware.objects.get(id=id)
    if request.method == 'POST':
        person = request.POST['person']
        if len(person) == 0:
            return redirect('index')
        Asignacion.objects.create(hardware=hardware, usuario=person)
        return redirect(reverse('index')+f'?asignacion=1')
    
    return redirect('index')


def asignaciones(request):
    ctx = {'link':'asignaciones'}
    asignaciones = Asignacion.objects.filter()
    
    
    ctx['asignaciones'] = asignaciones
    return render(request, 'main.html', ctx)

def importar_datos(request):
    wb = openpyxl.Workbook("nuevo_excel_t4.xlsx")

    wb.save('nuevo.xlsx')
    return redirect('index')